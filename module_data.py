import openpyxl
import os
from zipfile import ZipFile
import shutil
from bs4 import BeautifulSoup as bs
import re

import pandas
import pandas as pd
import time


class Macros:
    file_1c = ''
    file_m = ''
    sp_1c = None
    sp_m = None


    def convert_xlsx(self, file_in, new_filename):
        tmp_folder = 'tmp/convert_wrong_excel/'
        os.makedirs(tmp_folder, exist_ok=True)
        with ZipFile(file_in) as excel_container:
            excel_container.extractall(tmp_folder)
        wrong_file_path = os.path.join(tmp_folder, 'xl', 'SharedStrings.xml')
        correct_file_path = os.path.join(tmp_folder, 'xl', 'sharedStrings.xml')
        os.rename(wrong_file_path, correct_file_path)
        shutil.make_archive('tmp', 'zip', tmp_folder)
        os.replace('tmp.zip', new_filename)
        return True


    def get_sp_1c(self, filename, message_box):
        message_box.emit('читаем отчет 1с')
        self.file_1c = filename
        if self.convert_xlsx(self.file_1c, "list_pharmacy.xlsx") is False:
            message_box.emit('Ошибка конвертации справочника 1с с аптеками')
            return False
        wb = openpyxl.load_workbook("list_pharmacy.xlsx")
        sheet = wb.active
        for x in range(1, 20):
            color_h = sheet.cell(row=x, column=1).fill.fgColor.rgb
            if color_h == '00EBEBCC':
                skip_rows = x-1
                break
        self.sp_1c = pd.read_excel("list_pharmacy.xlsx", skiprows=skip_rows)
        self.sp_1c = self.sp_1c[["Заказчик", "Поставщик", "Соглашение (Кликабельно)"]]
        self.sp_1c = self.sp_1c.dropna(subset=['Поставщик', 'Заказчик'])
        os.remove('list_pharmacy.xlsx')
        message_box.emit('данные из 1с получены')

    def is_green_coll(self, coll):

        result = False
        style = coll.get('style')
        if style != None:
            p = re.compile(r"background-color: *lightgreen", re.I)
            if p.search(style):
                result = True
        return result


    def get_sp_monitoring(self, filename, message_box):
        message_box.emit('читаем отчет из мониторинга')
        self.sp_m = pd.DataFrame(columns=['Заказчик_мониторинг', 'Поставщик_мониторинг'])
        with open(filename, 'r', encoding='utf-8') as f:
            contents = f.read()
        html = bs(contents, 'lxml')
        rows = html.find_all('tr')
        obj_pharma = rows[0].find_all('td')
        list_pharma = []
        for ph in obj_pharma:
            list_pharma.append(ph.text.strip())
        for i in range(2, len(rows)):
            colls = rows[i].find_all('td')
            manuf = colls[0].text.strip()
            if manuf == '':
                continue
            for j in range(2, len(colls)):
                pharma = list_pharma[j]
                green = self.is_green_coll(colls[j])
                txt = colls[j].text
                if green is True:
                    self.sp_m = self.sp_m._append({'Заказчик_мониторинг': manuf, 'Поставщик_мониторинг': pharma},
                                              ignore_index=True)
        message_box.emit('данные из мониторинга получены')


    def compare_sp(self, list_pharma, message_box):
        sp_1c_copy = self.sp_1c
        sp_m_copy = self.sp_m
        file_out = 'compare.xlsx'
        try:
            with open(file_out, "w") as file:
                pass
        except IOError:
            metka = int(time.time())
            file_out = f'compare_{metka}.xlsx'
        if len(list_pharma) >0:
            sp_m_copy = sp_m_copy[sp_m_copy.Поставщик_мониторинг.isin(list_pharma)]
            sp_1c_copy = sp_1c_copy[sp_1c_copy.Поставщик.isin(list_pharma)]
        comp_search_1c = sp_m_copy.merge(sp_1c_copy , left_on = ['Заказчик_мониторинг','Поставщик_мониторинг'], right_on = ['Заказчик','Поставщик'], how = 'left')
        comp_search_1c = comp_search_1c[comp_search_1c['Поставщик'].isna()]
        comp_search_1c = comp_search_1c.drop(['Заказчик','Поставщик','Соглашение (Кликабельно)'],axis=1)
        comp_search_m = sp_1c_copy.merge(sp_m_copy, left_on = ['Заказчик','Поставщик'], right_on = ['Заказчик_мониторинг','Поставщик_мониторинг'],  how = 'left')
        comp_search_m = comp_search_m[comp_search_m['Заказчик_мониторинг'].isna()]
        comp_search_m = comp_search_m.drop(['Заказчик_мониторинг','Поставщик_мониторинг'], axis=1)
        writer = pd.ExcelWriter(file_out)
        comp_search_1c.to_excel(writer, 'Сравнение СП', index=False, startrow=1)
        comp_search_m.to_excel(writer, 'Сравнение СП', index=False, startrow=1, startcol=3)
        pharma_1c = self.sp_1c
        pharma_m = self.sp_m
        manuf_1c = self.sp_1c
        manuf_m = self.sp_m
        #сравнение поставщиков
        comp_1c_m = pharma_1c.merge(pharma_m, left_on='Поставщик', right_on='Поставщик_мониторинг', how='left')
        comp_1c_m = comp_1c_m[comp_1c_m['Поставщик_мониторинг'].isna()]
        comp_1c_m = pd.DataFrame(comp_1c_m.Поставщик.unique(), columns=['Поставщик отсутствует в мониторинге'])
        comp_1c_m.to_excel(writer, 'Сравнение наименований', index=False)

        comp_m_1c = pharma_m.merge(pharma_1c, left_on='Поставщик_мониторинг', right_on='Поставщик', how='left')
        comp_m_1c = comp_m_1c[comp_m_1c['Поставщик'].isna()]
        comp_m_1c = pd.DataFrame(comp_m_1c.Поставщик_мониторинг.unique(), columns=['Поставщик отсутствует в 1c'])
        comp_m_1c.to_excel(writer, 'Сравнение наименований', index=False, startcol=3)
        #сравнение заказчиков
        comp_1c_m = manuf_1c.merge(manuf_m, left_on='Заказчик', right_on='Заказчик_мониторинг', how='left')
        comp_1c_m = comp_1c_m[comp_1c_m['Заказчик_мониторинг'].isna()]
        comp_1c_m = pd.DataFrame(comp_1c_m.Заказчик.unique(), columns=['Заказчик отсутствует в мониторинге'])
        comp_1c_m.to_excel(writer, 'Сравнение наименований', index=False, startcol=6)
        comp_m_1c = manuf_m.merge(manuf_1c, left_on='Заказчик_мониторинг', right_on='Заказчик', how='left')
        comp_m_1c = comp_m_1c[comp_m_1c['Заказчик'].isna()]
        comp_m_1c = pd.DataFrame(comp_m_1c.Заказчик_мониторинг.unique(), columns=['Заказчик отсутствует в 1c'])
        comp_m_1c.to_excel(writer, 'Сравнение наименований', index=False, startcol=9)
        writer.close()
        self.decor_xlsx(file_out)
        message_box.emit(f'Файл сравнения создан {file_out}')


    def get_list_pharma(self):
        list_pharma = []
        if self.sp_1c is not None and self.sp_m is not None:
            list_1c = list(self.sp_1c.Поставщик.unique())
            list_m = list(self.sp_m.Поставщик_мониторинг.unique())
            list_pharma = sorted(list(set(list_1c + list_m)))
        return list_pharma

    def decor_xlsx(self, file_out):
        fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor='FFffff00')
        alignment = openpyxl.styles.Alignment(
            horizontal='center',
            vertical='center',
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=False,
            indent=0
        )
        wb = openpyxl.load_workbook(file_out)
        worksheet = wb['Сравнение СП']
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 30
        worksheet.column_dimensions['C'].width = 10
        worksheet.column_dimensions['D'].width = 30
        worksheet.column_dimensions['E'].width = 30
        worksheet.column_dimensions['F'].width = 20
        worksheet.merge_cells('A1:B1')
        worksheet['A1'] = 'Отсутствуют в 1С'
        worksheet['A1'].fill = fill
        worksheet['A1'].alignment = alignment
        worksheet.merge_cells('D1:F1')
        worksheet['D1'] = 'Отсутствуют в мониторинге'
        worksheet['D1'].fill = fill
        worksheet['D1'].alignment = alignment
        worksheet = wb['Сравнение наименований']
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['D'].width = 30
        worksheet.column_dimensions['G'].width = 30
        worksheet.column_dimensions['J'].width = 30
        wb.save(file_out)


if __name__ == "__main__":
    df = pd.read_html("1.html", encoding='utf-8')
    df = df[0]
    df.to_html("000.html", index=False)